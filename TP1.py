import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from funciones import calc_m, calc_b, calc_fc, calc_fd
from fisico_teorico import fisico_teorico
from cremer import cremer
from sharp import sharp
from davy import davy
from iso import iso

#definición de la clase 'Material'
class Material:
    def __init__(self, tipo, densidad, myoung, fperd, mpoisson, dimensiones):
        self.tipo = tipo
        self.rho = int(densidad)
        self.e = float(myoung)
        self.eta = float(fperd)
        self.sigma = float(mpoisson)
        self.dim = dimensiones
        self.m = calc_m(self.rho, self.dim)
        self.b = calc_b(self.e, self.sigma, self.dim)

#definición del array de frecuencias centrales y de los ticks para el gráfico
frecuencias = np.array([
    20, 25, 31.5, 40, 50, 63, 80, 100, 125, 160, 200, 250, 315, 400, 500, 630, 800, 1000,
    1250, 1600, 2000, 2500, 3150, 4000, 5000, 6300, 8000, 10000, 12500, 16000, 20000
])
freq_ticks = np.array([25, 50, 100, 200, 500, 1000, 2000, 4000, 8000, 16000])

#definición de constantes
c = 343
rho0 = 1.18

#lectura del archivo de materiales y definición del material y sus dimensiones
db = pd.read_excel('./materiales.xlsx')

#selección del material y su espesor; tamaño fijo en 3m * 7m
mats = db['Material'].tolist()
material_str = ''
for i, m in enumerate(mats):
    material_str += f'{i+1}. {m}.\n'
ind = int(input(f'Elija un material entre los siguientes:\n{material_str}Indice deseado: ' )) - 1
dim = (3, 7, float(input('Ingrese espesor del material (en metros): ')))

#instanciación de un objeto de la clase 'Material'
material = Material(db.iloc[ind, 1], db.iloc[ind, 2], db.iloc[ind, 3], db.iloc[ind, 4], db.iloc[ind, 5], dim)

#cálculo de la frecuencia crítica y de la frecuencia de densidad
fc = calc_fc(material.e, material.rho, material.dim, c)
fd = calc_fd(material.m, material.e, material.b, material.rho)

#cálculo de r con los modelos
r_fisico_teorico = fisico_teorico(frecuencias, material.m, material.eta, fc, fd, rho0, c)
r_iso = iso(frecuencias, material.m, material.eta, material.dim, fc, rho0, c)
r_cremer = cremer(frecuencias, material.m, material.eta, fc, fd)
r_sharp = sharp(frecuencias, material.m, material.eta, fc, rho0, c)
r_davy = davy(frecuencias, material.rho, material.e, material.sigma, material.dim, material.m, material.eta, fc, rho0, c)

#creación del array de resultados
data_R = [r_fisico_teorico, r_iso, r_cremer, r_sharp, r_davy]
modelos = ['Físico-teórico', 'ISO 12354-1', 'Cremer', 'Sharp', 'Davy']

res = pd.DataFrame(
    data = data_R,
    index = modelos,
    columns = frecuencias
)

#guardado en formato excel
with pd.ExcelWriter(f'./res/resultados_{material.tipo.lower()}_{material.dim[0]}x{material.dim[1]}x{str(material.dim[2]).replace(".", ",")}.xlsx', engine = 'openpyxl') as writer:
    res.to_excel(writer, sheet_name = 'R')
    worksheet = writer.sheets['R']

    relleno_encabezado = PatternFill(start_color = '1F497D', end_color = '1F497D', fill_type = 'solid')
    fuente_encabezado = Font(name = 'Calibri', size = 11, bold = True, color = 'FFFFFF')
    fuente_data = Font(name = 'Calibri', size = 11)
    fuente_modelos = Font(name = 'Calibri', size = 11, bold = True)
    borde_fino = Border(
        left = Side(style = 'thin', color = 'D9D9D9'),
        right = Side(style = 'thin', color = 'D9D9D9'),
        top = Side(style = 'thin', color = 'D9D9D9'),
        bottom = Side(style = 'thin', color = 'D9D9D9')
    )

    worksheet.cell(row = 1, column = 1).value = 'Modelo / Frecuencia (Hz)'
    worksheet.cell(row =1, column = 1).font = fuente_encabezado
    worksheet.cell(row = 1, column = 1).fill = relleno_encabezado
    worksheet.cell(row = 1, column = 1).alignment = Alignment(horizontal = "center", vertical = "center")
    worksheet.column_dimensions['A'].width = 24

    for col_idx in range(2, len(res.columns) + 2):
        cell = worksheet.cell(row = 1, column = col_idx)
        cell.font = fuente_encabezado
        cell.fill = relleno_encabezado
        cell.alignment = Alignment(horizontal = "center", vertical = "center")

        letra_col = cell.column_letter
        worksheet.column_dimensions[letra_col].width = 10
    
    for row_idx in range(2, len(res.index) + 2):
        worksheet.cell(row = row_idx, column = 1).font = fuente_modelos
        worksheet.cell(row = row_idx, column = 1).border = borde_fino

        for col_idx in range(2, len(res.columns) + 2):
            cell = worksheet.cell(row = row_idx, column = col_idx)
            cell.font = fuente_data
            cell.border = borde_fino
            cell.number_format = '0.0'
            cell.alignment = Alignment(horizontal = "right")     

#graficación de los resultados
plt.plot(frecuencias, r_fisico_teorico, label = 'Físico-teórico')
plt.plot(frecuencias, r_iso, label = 'ISO 12354-1')
plt.plot(frecuencias, r_cremer, label = 'Cremer')
plt.plot(frecuencias, r_sharp, label = 'Sharp')
plt.plot(frecuencias, r_davy, label = 'Davy')

#acondicionamiento del gráfico
plt.axvline(x=fc, color='black', ls='--', label = f'Frecuencia de coincidencia (≈{round(fc)} Hz)')
plt.ylim(bottom = 0, top = 130)
plt.xscale('log')
plt.xticks(frecuencias, labels = [str(f) for f in frecuencias], rotation = 45)
plt.grid(True, which = 'both', ls = '--', alpha = 0.5)
plt.legend()
plt.title(f'R para un panel simple de {material.tipo.lower()} de {material.dim[0]}m x {material.dim[1]}m x {str(material.dim[2]).replace(".", ",")}m', fontsize = '16')

plt.show()