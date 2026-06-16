import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.legend import Legend
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.layout import Layout, ManualLayout

def calc_m(rho, dim):
    m = rho * dim[2]
    return m

def calc_b(e, sigma, dim):
    b = (e * (dim[2] ** 3)) / (12 * (1 - sigma ** 2))
    return b

def calc_fc(e, rho, dim, c0 = 343):
    fc = ((c0 ** 2) / (1.8 * dim[2])) * np.sqrt(rho / e)
    return fc

def calc_fd(m, e, b, rho):
    fd = (e / (2 * np.pi * rho)) * np.sqrt(m / b)
    return fd

def calc_f11(fc, dim, c0 = 343):
    f11 = ((c0 ** 2) / (4 * fc)) * ((1 / (dim[0] ** 2)) + (1 / (dim[1] ** 2)))
    return f11

def calc_eta_total(f, m, eta):
    eta_total = eta + (m / (485 * np.sqrt(f)))
    return eta_total

def ley_masa(f, m):
    r = 20 * np.log10(m * f) - 47
    return r

def sigma(g, f, dim, c0 = 343):
    w = 1.3
    beta = 0.234
    n = 2

    s = dim[0] * dim[1]
    u = 2 * (dim[0] + dim[1])

    twoa = 4 * (s / u)

    k = 2 * np.pi * f / c0
    f_lim = w * np.sqrt(np.pi / (k * twoa))

    if f_lim > 1:
        f_lim = 1
    
    h = 1 / (np.sqrt(k * twoa / np.pi) * 2 / 3 - beta)
    q = 2 * np.pi / ((k ** 2) * s)
    qn = q ** n

    if g < f_lim:
        alpha = h / f_lim - 1
        xn = (h - alpha * g) ** n
    else:
        xn = g ** n
    
    rad = (xn + qn) ** (-1 / n)

    return rad

def shear(f, rho, e, sigma, dim):
    omega = 2 * np.pi * f

    chi = ((1 + sigma) / (0.87 + 1.12 * sigma)) ** 2

    x = (dim[2] ** 2) / 12
    qp = e / (1 - sigma ** 2)

    c = -omega * omega
    b = c * (1 + 2 * chi / (1 - sigma)) * x
    a = x * qp / rho

    kbcor2 = (-b + np.sqrt(b ** 2 - 4 * a * c)) / (2 * a)
    kb2 = np.sqrt(-c / a)

    g = e / (2 * (1 + sigma))

    kt2 = -c * rho * chi / g
    kl2 = -c * rho / qp
    ks2 = kt2 + kl2

    asi = (1 + x * (kbcor2 * kt2 / kl2 - kt2)) ** 2
    bsi = 1 - x * kt2 + kbcor2 * ks2 / (kb2 ** 2)
    csi = np.sqrt(1 - x * kt2 + (ks2 ** 2) / (4 * (kb2 ** 2)))

    out = asi / (bsi * csi)

    return out

def calc_radfree(f, fc, dim, c0 = 343):
    l1 = max(dim[0], dim[1])
    l2 = min(dim[0], dim[1])
    
    sig2 = 4 * l1 * l2 * ((f / c0) ** 2)
    sig3 = np.sqrt((2 * np.pi * f * (l1 + l2)) / (16 * c0))

    f11 = calc_f11(fc, dim, c0)

    if f11 <= (0.5 * fc):
        if f >= fc:    
            sig =  1 / (np.sqrt(1 - fc / f))
        else:
            lamb = np.sqrt(f / fc)
            
            d1 = ((1 - lamb ** 2) * np.log((1 + lamb) / (1 - lamb)) + 2 * lamb) / (4 * (np.pi ** 2) * ((1 - lamb ** 2) ** 1.5))
            d2 = 0 if f > (0.5 * fc) else (8 * (c0 ** 2) * (1 - 2 * (lamb ** 2))) / ((fc ** 2) * (np.pi ** 4) * l1 * l2 * lamb * np.sqrt(1 - (lamb ** 2)))
            
            sig = ((2 * (l1 + l2)) / (l1 * l2)) * (c0 / fc) * d1 + d2
    
        if f < f11 and f11 < (0.5 * fc) and sig > sig2:
            sig = sig2
    else:
        if f < fc and sig2 < sig3:
            sig = sig2
        elif f > fc and  (1 / (np.sqrt(1 - fc / f))) < sig3:
            sig =  1 / (np.sqrt(1 - fc / f))
        else:
            sig = sig3

    if sig > 2:
        sig = 2

    return sig

def calc_radforced(f, dim, c0 = 343):
    l1 = max(dim[0], dim[1])
    l2 = min(dim[0], dim[1])
    k0 = (2 * np.pi * f ) / c0

    lamb = -0.964 - (0.5 + l2 / (np.pi * l1)) * np.log(l2 / l1) + (5 * l2) / (2 * np.pi * l1) - 1 / (4 * np.pi * l1 * l2 * (k0 ** 2))

    sf = 0.5 * (np.log(k0 * np.sqrt(l1 * l2)) - lamb)

    if sf > 2:
        sf = 2

    return sf

def save_xlsx(res, material, modelos):
    mat_in_cols = ['Material', 'Densidad [kg/m³]', 'Módulo de Young [N/m²]', 'Factor de pérdidas', 'Módulo de Poisson', 'Dimensiones [m] x [m] x [m]']
    mat_in_ind = ['1']
    mat_in_data = [[material.tipo, material.rho, material.e, material.eta, material.sigma, ' x '.join([str(d) for d in material.dim])]]
    
    mat_in = pd.DataFrame(
                data = mat_in_data,
                index = mat_in_ind,
                columns = mat_in_cols
                )

    with pd.ExcelWriter(f'./res/resultados_{material.tipo.lower()}_{str(material.dim[0]).replace(".", ",")}x{str(material.dim[1]).replace(".", ",")}x{str(material.dim[2]).replace(".", ",")} ({", ".join(modelos)}).xlsx', engine = 'openpyxl') as writer:
        mat_in.to_excel(writer, sheet_name = 'R', index = False)
        res.to_excel(writer, sheet_name = 'R', startrow = 3)
        worksheet = writer.sheets['R']

        relleno_gris = PatternFill(start_color='696969', end_color='696969', fill_type='solid')
        relleno_gris_claro = PatternFill(start_color='BABABA', end_color='BABABA', fill_type='solid')
        fuente_encabezado_gris = Font(name='Calibri', size = 11, bold=True, color='FFFFFF')
        relleno_encabezado = PatternFill(start_color = '1F497D', end_color = '1F497D', fill_type = 'solid')
        fuente_encabezado = Font(name = 'Calibri', size = 11, bold = True, color = 'FFFFFF')
        fuente_data = Font(name = 'Calibri', size = 11)
        fuente_modelos = Font(name = 'Calibri', size = 11, bold = True)
        borde_fino = Border(
            left = Side(style = 'thin', color = 'D9D9D9'),
            right = Side(style = 'thin', color = 'D9D9D9'),
            top = Side(style = 'thin', color = 'D9D9D9'),
            bottom = Side(style = 'thin', color = 'D9D9D9')
        )

        for col_idx in range(1, len(mat_in.columns) + 1):
            cell = worksheet.cell(row = 1, column = col_idx)
            cell.font = fuente_encabezado_gris
            cell.fill = relleno_gris
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
            cell.border = borde_fino
            
        for col_idx in range(1, len(mat_in.columns) + 1):
            cell = worksheet.cell(row = 2, column = col_idx)
            cell.font = fuente_data
            cell.fill = relleno_gris_claro
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
            cell.border = borde_fino

        worksheet.cell(row = 4, column = 1).value = 'Modelo \ Frecuencia (Hz)'
        worksheet.cell(row = 4, column = 1).font = fuente_encabezado
        worksheet.cell(row = 4, column = 1).fill = relleno_encabezado
        worksheet.cell(row = 4, column = 1).alignment = Alignment(horizontal = 'center', vertical = 'center')

        for col_idx in range(2, len(res.columns) + 2):
            cell = worksheet.cell(row = 4, column = col_idx)
            cell.font = fuente_encabezado
            cell.fill = relleno_encabezado
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        
        for row_idx in range(2, len(res.index) + 2):
            worksheet.cell(row = row_idx + 3, column = 1).font = fuente_modelos
            worksheet.cell(row = row_idx + 3, column = 1).border = borde_fino

            for col_idx in range(2, len(res.columns) + 2):
                cell = worksheet.cell(row = row_idx + 3, column = col_idx)
                cell.font = fuente_data
                cell.border = borde_fino
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal = 'right')

        worksheet.row_dimensions[1].height = 36
        worksheet.row_dimensions[2].height = 36

        worksheet.column_dimensions['A'].width = 24

        col_fin_valores = len(res.columns) + 1

        for col_idx in range(2, col_fin_valores + 1):
            cell = worksheet.cell(row=4, column=col_idx)

            if cell.value is not None:
                cell.value = str(cell.value)

        colores = [
            '1F497D',
            'C00000',
            '76933C',
            '60497A',
            'E36C0A',
            '948A54',
            '31859C'
        ]

        grafico = LineChart()
        grafico.style = 13
        grafico.grouping = 'standard'

        grafico.layout = Layout(
            manualLayout = ManualLayout(
                x = 0.02,   # Mueve la cuadrícula a la derecha (da aire al eje Y: R [dB])
                y = 0.02,   # Baja la cuadrícula (da aire al título principal)
                w = 0.75,   # Achica el ancho de la cuadrícula (da espacio a la leyenda)
                h = 0.65    # Achica el alto de la cuadrícula (da aire abajo al eje X: Frecuencia)
            )
        )

        grafico.x_axis.axPos = 'b'
        grafico.x_axis.delete = False
        grafico.x_axis.title = 'Frecuencia [Hz]'
        grafico.x_axis.tickLblPos = 'low'
        grafico.x_axis.crosses = 'min'
        grafico.y_axis.scaling.min = 20
        grafico.y_axis.scaling.max = 20000
        grafico.x_axis.majorGridlines = ChartLines()

        grafico.y_axis.axPos = 'l'
        grafico.y_axis.delete = False
        grafico.y_axis.title = 'R [dB]'
        grafico.y_axis.tickLblPos = 'nextTo'
        grafico.y_axis.crosses = 'min'
        grafico.y_axis.scaling.min = 0
        grafico.y_axis.scaling.max = 120
        grafico.y_axis.majorUnit = 10
        grafico.y_axis.majorGridlines = ChartLines()

        xvalues = Reference(worksheet, min_col = 2, min_row = 4, max_col = col_fin_valores, max_row = 4)

        for i in range(len(res.index)):
            fila_actual = 5 + i
            yvalues = Reference(worksheet, min_col = 2, min_row = fila_actual, max_col = col_fin_valores, max_row = fila_actual)

            serie = Series(
                idx = i,
                order = i,
                val = NumDataSource(numRef = NumRef(f = yvalues)),
            )

            nombre_modelo = str(worksheet.cell(row = fila_actual, column = 1).value)
            serie.tx = SeriesLabel(v = nombre_modelo)

            serie.marker.symbol = 'none'
            serie.spPr.ln.solidFill = colores[i % len(colores)]
            serie.spPr.ln.w = 25000
            serie.smooth = True

            grafico.series.append(serie)

        grafico.set_categories(xvalues)

        if len(res.index) == 1:
            grafico.legend = None
            grafico.title = f'R (Modelo {str(res.index[0])}) - panel simple de {material.tipo.lower()} de {str(material.dim[0]).replace(".", ",")}m x {str(material.dim[1]).replace(".", ",")}m x {str(material.dim[2]).replace(".", ",")}m'
        else:
            grafico.legend = Legend(legendPos='r', overlay=False)

            mods = []
            for idx in res.index:
                mods.append(idx)
            
            grafico.title = f'R (Modelos {", ".join(mods)}) - panel simple de {material.tipo.lower()} de {str(material.dim[0]).replace(".", ",")}m x {str(material.dim[1]).replace(".", ",")}m x {str(material.dim[2]).replace(".", ",")}m'

        grafico.width = 20
        grafico.height = 16

        fila_grafico = len(res.index) + 8

        worksheet.add_chart(grafico, f'A{fila_grafico}')

def graficar(ax, f, r, nombre, color):
    ax.plot(f, r, label = f'{nombre}', color = color)
    ax.scatter(f, r, s = 20, color = color)